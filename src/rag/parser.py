import csv
import io
from typing import NamedTuple

import openpyxl
from docx import Document
from pypdf import PdfReader


class ExtractedSection(NamedTuple):
    content: str
    page_reference: str | None
    section_reference: str | None


def parse_document(file_content: bytes, filename: str) -> list[ExtractedSection]:
    """
    Parses uploaded file bytes into extracted text sections with page/section references.
    Supports PDF, DOCX, TXT, MD, CSV, and XLSX formats.
    """
    ext = filename.lower().split(".")[-1] if "." in filename else ""

    if ext == "pdf":
        return _parse_pdf(file_content)
    elif ext in ("docx", "doc"):
        return _parse_docx(file_content)
    elif ext in ("txt", "md"):
        return _parse_text(file_content)
    elif ext == "csv":
        return _parse_csv(file_content)
    elif ext in ("xlsx", "xls"):
        return _parse_xlsx(file_content)
    else:
        # Default text fallback
        return _parse_text(file_content)


def _parse_pdf(file_content: bytes) -> list[ExtractedSection]:
    sections: list[ExtractedSection] = []
    reader = PdfReader(io.BytesIO(file_content))
    for idx, page in enumerate(reader.pages, start=1):
        text = page.extract_text() or ""
        text = text.strip()
        if text:
            sections.append(
                ExtractedSection(
                    content=text,
                    page_reference=f"Page {idx}",
                    section_reference=None,
                )
            )
    return sections


def _parse_docx(file_content: bytes) -> list[ExtractedSection]:
    sections: list[ExtractedSection] = []
    doc = Document(io.BytesIO(file_content))
    current_heading: str | None = None
    current_text_lines: list[str] = []

    for para in doc.paragraphs:
        text = para.text.strip()
        if not text:
            continue
        if para.style and para.style.name.startswith("Heading"):
            if current_text_lines:
                sections.append(
                    ExtractedSection(
                        content="\n".join(current_text_lines),
                        page_reference=None,
                        section_reference=current_heading,
                    )
                )
                current_text_lines = []
            current_heading = text
        else:
            current_text_lines.append(text)

    if current_text_lines:
        sections.append(
            ExtractedSection(
                content="\n".join(current_text_lines),
                page_reference=None,
                section_reference=current_heading,
            )
        )

    return sections if sections else [ExtractedSection(content="", page_reference=None, section_reference=None)]


def _parse_text(file_content: bytes) -> list[ExtractedSection]:
    text = file_content.decode("utf-8", errors="replace").strip()
    return [ExtractedSection(content=text, page_reference=None, section_reference=None)]


def _parse_csv(file_content: bytes) -> list[ExtractedSection]:
    text_content = file_content.decode("utf-8", errors="replace")
    reader = csv.reader(io.StringIO(text_content))
    lines: list[str] = []
    for row in reader:
        if row:
            lines.append(" | ".join(row))
    full_text = "\n".join(lines)
    return [ExtractedSection(content=full_text, page_reference=None, section_reference=None)]


def _parse_xlsx(file_content: bytes) -> list[ExtractedSection]:
    sections: list[ExtractedSection] = []
    wb = openpyxl.load_workbook(io.BytesIO(file_content), data_only=True)
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        lines: list[str] = []
        for row in sheet.iter_rows(values_only=True):
            row_vals = [str(cell) for cell in row if cell is not None]
            if row_vals:
                lines.append(" | ".join(row_vals))
        if lines:
            sections.append(
                ExtractedSection(
                    content="\n".join(lines),
                    page_reference=f"Sheet: {sheet_name}",
                    section_reference=sheet_name,
                )
            )
    return sections
